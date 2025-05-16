from openpyxl import Workbook

class ExcelReport:
    def __init__(self, filename):
        self.filename = filename
        self.wb       = Workbook()
        self.ws       = self.wb.active

    def add_title(self, title, row=1, col=1):
        self.ws.cell(row=row, column=col, value=title)

    def add_rows(self, rows, start_row=2):
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, val in enumerate(row, start=1):
                self.ws.cell(row=r_idx, column=c_idx, value=val)

    def save(self):
        self.wb.save(self.filename)
